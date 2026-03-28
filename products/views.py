from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, TemplateView
from django.db.models import Count, Q
from django.db.models.aggregates import Avg
# from .tasks import send_emails
from .models import Product, Brand, Review, ProductImages, Category, Offer, Bundle
# Create your views here.
from django.http import JsonResponse, HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator
from django.contrib import messages
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter



class BundleList(ListView):
    model = Bundle
    queryset = Bundle.objects.all()
    template_name = 'products/bundle_list.html'
    paginate_by = 20


class ProductList(ListView):
    model = Product
    paginate_by = 20

    def get_queryset(self):
        queryset = Product.objects.filter(product_type='single')
        query = self.request.GET.get('q')
        if query:
            queryset = queryset.filter(
                Q(name__icontains=query) | 
                Q(descriptions__icontains=query)
            )
        return queryset
    



class ProductDetail(DetailView):
    model =Product
    queryset = Product.objects.all()



    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["reviews"] = Review.objects.filter(product=self.get_object())
        context["related_products"] = Product.objects.filter(brand=self.get_object().brand) 
        return context
    

class BrandList(ListView):
    model = Brand    #context : object_list, model_list
    # paginate_by = 20 # User wants all brands and products on one page
    queryset = Brand.objects.prefetch_related('product_brand').all()
 

class BrandDetail(ListView):
    model = Product     #context : object_list, model_list
    template_name = 'products/brand_detail.html'


    def get_queryset(self):
        brand = get_object_or_404(Brand, slug=self.kwargs['slug'])
        return super().get_queryset().filter(brand=brand) 
    

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["brand"] = get_object_or_404(Brand, slug=self.kwargs['slug'])
        return context


class CategoryList(ListView):
    model = Category
    paginate_by = 20
    queryset = Category.objects.all()


class CategoryDetail(ListView):
    model = Product     #context : object_list, model_list
    template_name = 'products/category_detail.html'

    def get_queryset(self):
        category = get_object_or_404(Category, slug=self.kwargs['slug'])
        return super().get_queryset().filter(category=category)  
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["category"] = get_object_or_404(Category, slug=self.kwargs['slug'])
        return context


class OfferList(ListView):
    model = Offer
    queryset = Offer.objects.all()
    paginate_by = 20


class OfferDetail(DetailView):
    model = Offer
    queryset = Offer.objects.all()


@user_passes_test(lambda u: u.is_staff)
def export_products_excel(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    # Header
    columns = ['Name', 'SKU', 'Old Price', 'New Price', 'Quantity', 'Brand', 'Categories', 'Flag', 'Type']
    ws.append(columns)

    for product in Product.objects.all():
        brand_name = product.brand.name if product.brand else ""
        categories = ", ".join([cat.name for cat in product.category.all()])
        ws.append([
            product.name,
            product.sku or "",
            product.old_price,
            product.new_price,
            product.quantity,
            brand_name,
            categories,
            product.flag,
            product.product_type
        ])

    wb.save(response)
    return response


@user_passes_test(lambda u: u.is_staff)
def download_product_template(request):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="products_template.xlsx"'

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"

    # Header
    columns = ['Name', 'SKU', 'Old Price', 'New Price', 'Quantity', 'Brand', 'Categories', 'Flag', 'Type']
    ws.append(columns)
    
    # Add an example row
    ws.append(['Product Name', 'SKU123', 100, 80, 50, 'Brand Name', 'Category1, Category2', 'sale', 'single'])

    wb.save(response)
    return response


@user_passes_test(lambda u: u.is_staff)
def import_products_excel(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Skip header
            rows = list(ws.iter_rows(values_only=True))[1:]
            
            imported_count = 0
            updated_count = 0
            
            for row in rows:
                if not row[0]: continue # Skip empty rows
                
                name = row[0]
                sku = str(row[1]) if row[1] else None
                old_price = float(row[2]) if row[2] else 0
                new_price = float(row[3]) if row[3] else 0
                quantity = int(row[4]) if row[4] else 0
                brand_name = row[5]
                categories_str = row[6]
                flag = row[7] or 'new'
                product_type = row[8] or 'single'
                
                # Brand lookup
                brand = None
                if brand_name:
                    brand, _ = Brand.objects.get_or_create(name=brand_name)
                
                # Product lookup (by SKU or Name)
                product = None
                if sku:
                    product = Product.objects.filter(sku=sku).first()
                if not product:
                    product = Product.objects.filter(name=name).first()
                
                if product:
                    product.name = name
                    product.old_price = old_price
                    product.new_price = new_price
                    product.quantity = quantity
                    product.brand = brand
                    product.flag = flag
                    product.product_type = product_type
                    product.save()
                    updated_count += 1
                else:
                    product = Product.objects.create(
                        name=name,
                        sku=sku,
                        old_price=old_price,
                        new_price=new_price,
                        quantity=quantity,
                        brand=brand,
                        flag=flag,
                        product_type=product_type
                    )
                    imported_count += 1
                
                # Categories
                if categories_str:
                    product.category.clear()
                    cat_names = [c.strip() for c in str(categories_str).split(',')]
                    for cat_name in cat_names:
                        cat, _ = Category.objects.get_or_create(name=cat_name)
                        product.category.add(cat)
            
            messages.success(request, f"تم بنجاح: استيراد {imported_count} منتج جديد وتحديث {updated_count} منتج.")
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء الاستيراد: {str(e)}")
            
    next_url = request.GET.get('next') or request.POST.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('products:product_list')



